import openpyxl
from openpyxl import *
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

#Intro Sheets

#overall doc to scan (gives user choice)
FinanceSource_file = load_workbook('insert file')
#overall sheet to scan
FinanceSheet = FinanceSource_file['insert file']
#doc for LTSC
DatabaseSource_file = load_workbook("insert file")
#sheet for LTSC
DatabaseSheet = DatabaseSource_file["insert file"]
#doc for Windows 10 pro
Win10ProSource_file = load_workbook("insert file")
#sheet for windows 10 pro
Win10ProSheet = Win10ProSource_file["insert file"]

#sheet of all assets in lansweeper for use in finding out if it has been connected in 90 days
LastSeenDBSource_file = load_workbook("insert file")

#sheet for all assets
DBSheet = LastSeenDBSource_file["report"]



#####################################################################################################################################################################
#CHECK IF IT RUNS LTSC OR WINDOWS 10 PRO:
#Get the item to look for using loop(financeserialnumber) through main document (named Financesheet)
for i in range(1,3000):

    financeserialnumber = FinanceSheet.cell(row = i, column = 2).value

    #in case it is a blank column and skips the iteration if it is so the script does not error
    if financeserialnumber == None:
    
        continue
    
    else:

        #check if it has windows 10 pro by looping through database excel sheet of computers running WIN10 Pro
        for l in range(1,50):

            win10check = Win10ProSheet.cell(row = l, column = 1).value

            #checks if it is blank so that the script does not error out and skips the iteration if it is blank
            if win10check == None:
                continue
            
            else:
            
                LinString = 'AI'+ str(l)
                #checks each cell being looped through using string.find
                
                valueoferror = win10check.find(financeserialnumber)

                #if it is not found skip the cell
                if valueoferror == -1:
                    continue

                #string.find assigns a negative value if not found so if positive it means its there.
                # Assigns the cell spot created above with marking noting it does have windows 10 pro 
                else:
                    
                    FinanceSheet[LinString] = 'YES'

        #Check if it runs LTSC
        for j in range(1,400):

            #Checks through entire database of PCs running LTSC using loop(databaseserialnumber) (document named databaseSHEET)
            databaseserialnumber = DatabaseSheet.cell(row = j, column = 1).value
            
            #if the cell is blank it skips over it
            if databaseserialnumber == None:
                continue

            #Keeps looking if not blank
            else:
                
                #assigns spot to put marker in column AI in case it i found
                finalstring = 'AI' + str(i)
                
                #uses string.find() function to look for the value in scanned sheet and searches 
                #for it in the currently iterated cell of the LTSC database sheet
                valueoferror = databaseserialnumber.find(financeserialnumber)
                
                #assigns it -1 f not found and skips over the iteration
                if valueoferror == -1:
                    continue
                    
                #otherwise it keeps looking
                else:

                    #if the values match in the scanned sheet and the database sheet than it TAGS it in the selected spot YES
                    FinanceSheet[finalstring] = 'YES'
#################################################################################################################################################################
        #CHECKS AGAIN FOR ASSET NAME INCASE ASSET NAME IS NOT SERIAL NUMBER
        #loops through the entire lansweeper database
        for looper in range(1,4000):

                    #gets value of each cell in loop
                    lastSN = DBSheet.cell(row = looper, column = 29).value

                    #prevents crash if cell is blank
                    if lastSN == None:
                        continue
                    
                    else:
                        #checks if the serial number from finance match the serial number from the database
                        SNMatch = lastSN.find(financeserialnumber)

                        #If it doesnt move on to next
                        if SNMatch == -1:
                            continue

                        #if the Database contains the serial number from finance report  
                        else:
                            
                            #grab the corresponding AssetName value
                            AssetName = DBSheet.cell(row = looper, column = 1).value
                            
                            #start loop to scan LSTC DB for a match
                            for loopLSTCDB in range(1,400):
                                
                                #get value of each cell
                                AssetNameCheck = DatabaseSheet.cell(row = loopLSTCDB, column = 1).value
                                
                                #if cell is blank asset has no name so move on
                                if AssetNameCheck == None:
                                    continue
                                
                                #if cell is not blank check if it matches the asset name
                                else:
                                    AssetCross = AssetNameCheck.find(AssetName)
                                    
                                    #if the value does not match move on to next cell
                                    if AssetCross == -1:
                                        continue

                                    #if it does match then flag it.
                                    else:
                                        Crossfinalstring = 'AI' + str(i)

                                        FinanceSheet[Crossfinalstring] = 'YES'

#closes the dtabase files for LTSC and windows 10 pro as to not bog down the memory and processing speeds.
DatabaseSource_file.close()
Win10ProSource_file.close()

####################################################################################################################################
#FLAGS ON LAST SEEN DATE

#iterates though document
for p in range(1,3000):
    thefinancenumber = FinanceSheet.cell(row = p, column = 2).value

    #make sure that there is value in the cell so it doesnt crash
    if thefinancenumber == None:
        continue

    #If it has value count through the Database to see if it has been connected
    else:

        #Starts Loop to look through database sheet
        for q in range(1,4000):

            #checks serial number column as variable DBserial
            DBserial = DBSheet.cell(row = q, column = 29).value

            #ensures that the cell does not equal none so that the script does not crash
            if DBserial == None:
                continue

            #if it is not none it keeps searching
            else:
                
                #assigns spot to put marker
                qinString = 'AJ'+ str(p)
                
                #it checks through to search each cell being iterated through to see if it contains the value from the scanned sheet
                lastseenvalueoferror = DBserial.find(thefinancenumber)

                #if it doesnt match it returns the value -1 and if it is -1 then it skips over the iteration
                if lastseenvalueoferror == -1:
                    continue
                    
                #if it does match it assigns the value YES to the cell
                else:
                    
                    FinanceSheet[qinString] = 'YES'

#closes the file to make the program run more efficiently
LastSeenDBSource_file.close()

########################################################################################################################

#LAPTOP CHECK/ FLAG ENTRY FOR OS AND CONNECTION

#makes sure its a laptop or computer so that the OS flag applies to it
for k in range(1,3000):
    
    #checks through all cells of the scanned sheets descriptions column
    nameGet = FinanceSheet.cell(row = k, column = 4).value
    
    #uses strin.find functions to check if laptop or computer is in the description
    #if not a blank cell because a blank cell would cause a crash
    if nameGet != None:
        laptop = nameGet.find("aptop")
        computer = nameGet.find("omputer")

    #if it is a laptop or computer
    if laptop != -1 or computer != -1:
        
        #checks the columns of OS marks
        OScorrection = FinanceSheet.cell(row = k, column = 35).value
        
        #checks if it has a YES or NO
        if OScorrection == None:

            #Assigns the flag column with the flag O
            KinString = 'AH' + str(k) 
            FinanceSheet[KinString] = 'O,'

#################################################################################################################################################################
#CHECKS LAST SEEN DATE

#loops through scanned documents description column
for x in range(1,3000):
    
    xnameGet = FinanceSheet.cell(row = x, column = 4).value
    
    #checks if laptop and computer are in the title
    #Ensure that the cell isnt blank so that the script doesnt error out
    if xnameGet != None:

        xlaptop = xnameGet.find("aptop")

        xcomputer = xnameGet.find("omputer")

    if xlaptop != -1 or xcomputer != -1:
        
        #loops through the column containing the flag markings for the YES for itr has been seen on the database
        Lastcorrection = FinanceSheet.cell(row = x, column = 36).value
        
        #If it has not it assigns the flag D in the AH column
        if Lastcorrection == None:

            xKinString = 'AH' + str(x)

            oldvalue = FinanceSheet[xKinString].value

            #checks if there are other flags there as to not overwrite them
            if oldvalue == None:

                FinanceSheet[xKinString] = 'D,'

            else:
                #if there are other flags it gets those and adds the flag D to them
                FinanceSheet[xKinString] = FinanceSheet[xKinString].value + 'D,'         

##########################################################################################################################################################################
   
#DATE CHECK / FLAG

#checks Date column of scanned sheet through loop (named date)
for b in range(2,3000):
    date = FinanceSheet.cell(row = b, column = 9).value

#if the cell id is blank than it flags it as out of date
    if date == None: 

        #assigns it to flag column AH
        columnAH = 'AH' + str(b)

        #checks if the flag has a value already, if not it just assigns the flag
        if FinanceSheet[columnAH].value == None:
            FinanceSheet[columnAH].value = 'C,'

        #If it does have a value it retrieves the value and adds the new flag C on top of it
        else:
            FinanceSheet[columnAH].value = FinanceSheet[columnAH].value + 'C,'


    #when the cell is not blank it checks the date listed in the cell
    else:

        #converts the date to a string
        dateasstring = str(date)

        #seperates the year of the string and sets the type as an integer for operator use
        dateyear = int(dateasstring[0:4])

        #If less than 2019 than it is flagged
        if dateyear < 2019:

            #Assigns use in flag column AH
            columnAH = 'AH' + str(b)

            #If no other flags exist it flags it with C
            if FinanceSheet[columnAH].value == None:
                FinanceSheet[columnAH].value = 'C,'

            #if other values exist it  retreives them and flags it with C
            else:
                FinanceSheet[columnAH].value = FinanceSheet[columnAH].value + 'C,'

#sets up column headers for flags for Analysts SA
FinanceSheet['AH1'] = 'FLAGS' 
FinanceSheet['AI1'] = 'LTSC or Windows 10?'                   
FinanceSheet['AJ1'] = 'Seen In 3 mo'

#Saves the email
FinanceSource_file.save('insert file')   
FinanceSource_file.close()